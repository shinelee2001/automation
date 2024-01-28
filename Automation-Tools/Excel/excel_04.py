import openpyxl as excel
import shodan

api_key = "my_key"

api = shodan.Shodan(api_key)

book = excel.Workbook()
sheet = book.active

try:
        # Search Shodan
        results = api.search('tomcat', page=1)

        # Show the results
        print('Results found: {}'.format(results['total']))
        i=1
        for result in results['matches']:
                cell_ip = sheet.cell(row=i,column=1)
                cell_ip.value = result['ip_str']
                
                cell_org = sheet.cell(row=i, column=2)
                cell_org.value = result['org']
                i += 1
except shodan.APIError as e:
        err = 'Error: {}'.format(e)
        print(err)
        sheet.cell(row=1, column=1, value=str(err))

sheet.column_dimensions['A'].width = 32
book.save("py_excel03.xlsx")