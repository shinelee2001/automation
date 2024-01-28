import openpyxl as excel
book = excel.load_workbook("py_excel01.xlsx")
sheet = book.worksheets[0]

# cell 메소드 이용한 값입력 예제 1
sheet.cell(row=2, column=1, value="행렬을 이용하여 값 입력")

# cell 메소드 이용한 값입력 예제 2
cell = sheet.cell(row=3, column=1)
cell.value = "행렬을 이용하여 값 입력2"

book.save("py_excel02.xlsx")